#!/usr/bin/env python3
"""D1 State 全景 Excel — KVBt MT4 风格。

每只股票 3 行（今天/昨天/前天），每行横排：
  代码│名称│行业│日期│收盘│MN1│W1│D1│ef│
  主力净(T)│大单(T)│主力%(T)│主力净(Y)│大单(Y)│主力%(Y)│主力净(P)│大单(P)│主力%(P)│
  ADX│ATR%│MN1支│MN1阻│W1支│W1阻│D1支│D1阻

色彩：Hex 列蓝=E/F 绿=正 黄=0-7 红=负。ef 列蓝≥3 绿≥2。
"""

from collections import defaultdict
from pathlib import Path

import duckdb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/lv111101/Documents/hermass-observer-product")
MF_DB = ROOT / "outputs" / "blackwolf_moneyflow" / "blackwolf_moneyflow.duckdb"


def _load_3day_state():
    dbs = sorted(ROOT.glob("outputs/p116_foundation_20*/p116_foundation.duckdb"))
    if not dbs:
        dbs = sorted(ROOT.glob("outputs/p116_foundation_*/p116_foundation.duckdb"))
    con = duckdb.connect(str(dbs[-1]), read_only=True)
    dates = [
        str(d[0])
        for d in con.execute(
            "SELECT DISTINCT state_date FROM d1_perspective_state ORDER BY state_date DESC LIMIT 3"
        ).fetchall()
    ]
    rows = con.execute(
        f"""
        SELECT state_date, stock_code, d1_close,
               mn1_state_hex, w1_state_hex, d1_state_hex,
               mn1_state_score, w1_state_score, d1_state_score, ef_count,
               d1_adx14, d1_atr_ratio_pct,
               d1_sr_support, d1_sr_resistance,
               w1_sr_support, w1_sr_resistance,
               mn1_sr_support, mn1_sr_resistance
        FROM d1_perspective_state
        WHERE state_date IN ({",".join("?" * len(dates))})
        ORDER BY state_date DESC, ef_count DESC
    """,
        dates,
    ).fetchall()
    con.close()
    out = defaultdict(dict)
    for r in rows:
        ds, code = str(r[0]), r[1]
        out[ds][code] = {
            "close": r[2],
            "hex": (r[3], r[4], r[5]),
            "sc": (r[6], r[7], r[8]),
            "ef": r[9],
            "adx": r[10],
            "atr": r[11],
            "sr": {"d": (r[12], r[13]), "w": (r[14], r[15]), "m": (r[16], r[17])},
        }
    return {"dates": dates, "data": dict(out)}


def _load_moneyflow():
    if not MF_DB.exists():
        return {}
    con = duckdb.connect(str(MF_DB), read_only=True)
    dates = [
        str(d[0])
        for d in con.execute(
            "SELECT DISTINCT date FROM moneyflow_daily ORDER BY date DESC LIMIT 3"
        ).fetchall()
    ]
    rows = con.execute(
        f"""
        SELECT stock_code, date, active_net, big_order_net, active_net_ratio
        FROM moneyflow_daily WHERE date IN ({",".join("?" * len(dates))})
    """,
        dates,
    ).fetchall()
    con.close()
    out = defaultdict(dict)
    for code, d, an, bo, r in rows:
        c6 = code.split(".")[0] if "." in code else code[-6:]
        out[c6][str(d)] = {
            "an": round(an / 1e8, 2) if an else 0,
            "bo": round(bo / 1e8, 2) if bo else 0,
            "rt": round(r * 100, 1) if r else 0,
        }
    return {"dates": dates, "data": dict(out)}


def _hex_fill(hex_val) -> PatternFill:
    h = str(hex_val)
    if h in ("E", "F"):
        return PatternFill("solid", fgColor="CCE5FF")
    if h.startswith("-"):
        return PatternFill("solid", fgColor="F8D7DA")
    try:
        s = int(h, 16)
        if s >= 8:
            return PatternFill("solid", fgColor="D4EDDA")
        return PatternFill("solid", fgColor="FFF3CD")
    except:
        return PatternFill("solid", fgColor="FFF3CD")


def generate():
    s3 = _load_3day_state()
    dates = s3["dates"]
    sd = s3["data"]
    latest = dates[0]
    codes = sorted(sd.get(latest, {}).keys())
    mf = _load_moneyflow()
    mfd = mf.get("dates", [])
    mfv = mf.get("data", {})

    nm, sw = {}, {}
    fdb = ROOT / "outputs/fundamental/fundamental_evidence.duckdb"
    if fdb.exists():
        c = duckdb.connect(str(fdb), read_only=True)
        for code, name, sw1 in c.execute(
            "SELECT stock_code,stock_name,sw_l1 FROM ifind_industry_chain_profile"
        ).fetchall():
            nm[code] = name or ""
            sw[code] = sw1 or ""
        c.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"State_{latest}"

    T = mfd[0] if len(mfd) > 0 else "T"
    Y = mfd[1] if len(mfd) > 1 else "Y"
    P = mfd[2] if len(mfd) > 2 else "P"

    headers = [
        "代码",
        "名称",
        "行业",
        "日期",
        "收盘",
        "MN1",
        "W1",
        "D1",
        "ef",
        f"主力净({T})",
        f"大单({T})",
        f"主力%({T})",
        f"主力净({Y})",
        f"大单({Y})",
        f"主力%({Y})",
        f"主力净({P})",
        f"大单({P})",
        f"主力%({P})",
        "ADX",
        "ATR%",
        "MN1支",
        "MN1阻",
        "W1支",
        "W1阻",
        "D1支",
        "D1阻",
    ]

    hf = Font(bold=True, size=10, color="FFFFFF")
    fill_hdr = PatternFill("solid", fgColor="2F5496")
    thin = Side(style="thin", color="E0E0E0")
    bdr = Border(bottom=thin)
    cen = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hf
        cell.fill = fill_hdr
        cell.alignment = cen

    row = 2
    for code in codes:
        nc = nm.get(code, "")
        sc = sw.get(code, "")
        for di, d in enumerate(dates):
            s = sd.get(d, {}).get(code)
            if s is None:
                continue
            c6 = code.split(".")[0] if "." in code else code[-6:]
            mfr = mfv.get(c6, {})

            # KVBt-style: time label
            day_label = f"{d} (今)" if di == 0 else (f"{d} (昨)" if di == 1 else f"{d} (前)")

            vals = [
                code,
                nc,
                sc,
                day_label,
                round(s["close"], 2),
                s["hex"][0],
                s["hex"][1],
                s["hex"][2],
                s["ef"],
            ]
            for mdd in mfd:
                m = mfr.get(mdd, {})
                vals.extend([m.get("an", ""), m.get("bo", ""), m.get("rt", "")])
            vals.append(round(s["adx"], 1) if s["adx"] else "")
            vals.append(round(s["atr"], 2) if s["atr"] else "")
            vals.extend(
                [
                    s["sr"]["m"][0] if s["sr"]["m"][0] else "",
                    s["sr"]["m"][1] if s["sr"]["m"][1] else "",
                    s["sr"]["w"][0] if s["sr"]["w"][0] else "",
                    s["sr"]["w"][1] if s["sr"]["w"][1] else "",
                    s["sr"]["d"][0] if s["sr"]["d"][0] else "",
                    s["sr"]["d"][1] if s["sr"]["d"][1] else "",
                ]
            )

            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.alignment = cen
                cell.border = bdr

            # Color MN1/W1/D1 hex columns (6,7,8)
            for pi in range(3):
                ws.cell(row=row, column=6 + pi).fill = _hex_fill(s["hex"][pi])
            # Color ef column
            ef = s["ef"]
            if ef >= 3:
                ws.cell(row=row, column=9).fill = PatternFill("solid", fgColor="CCE5FF")
            elif ef >= 2:
                ws.cell(row=row, column=9).fill = PatternFill("solid", fgColor="D4EDDA")

            # Color moneyflow: pink=net buy, green=net sell
            for mdi in range(len(mfd)):
                an = vals[10 + mdi * 3] if 10 + mdi * 3 <= len(vals) else None
                if an and isinstance(an, (int, float)) and float(an) > 0:
                    ws.cell(row=row, column=10 + mdi * 3).fill = PatternFill("solid", fgColor="FFD6D6")
                elif an and isinstance(an, (int, float)) and float(an) < 0:
                    ws.cell(row=row, column=10 + mdi * 3).fill = PatternFill("solid", fgColor="D6FFD6")

            row += 1

    widths = [14, 12, 10, 18, 10, 7, 7, 7, 6, 12, 12, 11, 12, 12, 11, 12, 12, 11, 8, 8, 9, 9, 9, 9, 9, 9]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    ws.freeze_panes = "F2"

    ymd = latest.replace("-", "")
    out_dir = ROOT / "outputs" / "daily_state_excel"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"state_full_{ymd}.xlsx"
    wb.save(str(path))
    return path


if __name__ == "__main__":
    import json

    p = generate()
    print(
        json.dumps(
            {"status": "ok", "path": str(p), "size_kb": round(p.stat().st_size / 1024, 1)}, ensure_ascii=False
        )
    )
